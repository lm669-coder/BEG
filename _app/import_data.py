"""
Import flexible des données Excel vers SQLite.
Accepte n'importe quel fichier .xlsx/.xlsm avec sélection de feuille.
Détecte les colonnes par nom d'en-tête (insensible à la casse/accents).
"""

import io
from datetime import datetime

import openpyxl


# ── Helpers ───────────────────────────────────────────────────────────────────
def _safe(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    return (
        s
        if s and s.lower() not in ("none", "nan", "#n/a", "#ref!", "#value!")
        else None
    )


def _num(val):
    try:
        v = str(val).replace(",", ".").replace(" ", "").replace("\xa0", "")
        return float(v) if v else None
    except (ValueError, TypeError):
        return None


_ACCENT_MAP = {
    "é": "e",
    "è": "e",
    "ê": "e",
    "ë": "e",
    "à": "a",
    "â": "a",
    "ä": "a",
    "ù": "u",
    "û": "u",
    "ü": "u",
    "î": "i",
    "ï": "i",
    "ô": "o",
    "ö": "o",
    "ç": "c",
}


def _normalize(text):
    """Lowercase + strip accents for fuzzy header matching."""
    if not text:
        return ""
    t = str(text).lower().strip()
    for k, v in _ACCENT_MAP.items():
        t = t.replace(k, v)
    return t


def _find_col(headers_norm, *keywords):
    """Return first column index whose normalised header contains any keyword."""
    for i, h in enumerate(headers_norm):
        for kw in keywords:
            if kw in h:
                return i
    return None


# ── Lecture fichier ───────────────────────────────────────────────────────────
def open_workbook(source):
    """source = path string OR file-like object (st.file_uploader bytes)."""
    if hasattr(source, "read"):
        data = source.read()
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    return openpyxl.load_workbook(source, read_only=True, data_only=True)


def get_sheet_names(source):
    wb = open_workbook(source)
    names = wb.sheetnames
    wb.close()
    return names


def get_sheet_preview(source, sheet_name, max_rows=4):
    """Returns list of rows (list of cell values) for preview."""
    wb = open_workbook(source)
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i >= max_rows:
            break
        rows.append(list(row))
    wb.close()
    return rows


# ── Import chantiers ──────────────────────────────────────────────────────────
def import_chantiers(
    source, sheet_name, conn, header_row=1, first_data_row=2, overwrite=False
):
    """
    Import depuis n'importe quelle feuille.
    Détecte les colonnes via les en-têtes.
    Colonnes attendues (par mots-clés) :
      id_chantier : 'id', 'num', 'ref', 'chantier'
      gestionnaire : 'gest'
      intitule : 'intitul', 'nom'
      secteur, adresse, province, distance, client, montant, ...
    """
    wb = open_workbook(source)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    wb.close()

    if not rows:
        return 0, 0, "Feuille vide"

    headers_raw = [str(c) if c is not None else "" for c in rows[0]]
    headers_norm = [_normalize(h) for h in headers_raw]

    # Détection colonnes
    col = {
        "id": _find_col(
            headers_norm, "id chantier", "num chantier", "reference", " id", "^id"
        ),
        "dep": _find_col(headers_norm, "dep"),
        "gest": _find_col(headers_norm, "gest"),
        "intitule": _find_col(headers_norm, "intitul", "designation", "libelle"),
        "secteur": _find_col(headers_norm, "secteur"),
        "adresse": _find_col(headers_norm, "adresse"),
        "province": _find_col(headers_norm, "province"),
        "distance": _find_col(headers_norm, "distance"),
        "client": _find_col(headers_norm, "client"),
        "type_client": _find_col(headers_norm, "type client", "type_client"),
        "montant": _find_col(headers_norm, "montant"),
        "delai": _find_col(headers_norm, "delai exec"),
        "date_marche": _find_col(headers_norm, "marche gagne", "marche"),
        "ordre": _find_col(headers_norm, "ordre", "commencer"),
        "rp_dem": _find_col(headers_norm, "rp dem"),
        "rp_real": _find_col(headers_norm, "rp real"),
        "rd_dem": _find_col(headers_norm, "rd a dem", "rd dem"),
        "rd_real": _find_col(headers_norm, "rd real"),
        "rp_statut": _find_col(headers_norm, "rp - statut", "rp statut"),
        "rd_statut": _find_col(headers_norm, "rd - statut", "rd statut"),
        "n_caut": _find_col(headers_norm, "cautionnement", "n° caut"),
        "organisme": _find_col(headers_norm, "organisme"),
        "mont_caut": _find_col(headers_norm, "montant caut"),
        "doc_rp": _find_col(headers_norm, "document rp"),
        "doc_rd": _find_col(headers_norm, "document rd"),
        "pr": _find_col(headers_norm, "prix de revient", "prix revient", "pr soum"),
    }

    # Fallback PR : correspondance exacte sur "pr" (évite de matcher "province" etc.)
    if col["pr"] is None:
        col["pr"] = next(
            (i for i, h in enumerate(headers_norm) if h.strip() == "pr"), None
        )

    # Fallback: si aucun header "id chantier" trouvé, essayer colonne A
    if col["id"] is None:
        col["id"] = 0

    def g(row, key):
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    if overwrite:
        conn.execute("DELETE FROM chantiers")
        conn.commit()

    existing_ids: set[str] = set()
    if not overwrite:
        existing_ids = {
            r[0] for r in conn.execute("SELECT id_chantier FROM chantiers").fetchall()
        }

    imported = skipped = 0
    for row in rows[first_data_row - header_row :]:
        id_c = _safe(g(row, "id"))
        if not id_c:
            continue
        if id_c in existing_ids:
            skipped += 1
            continue

        conn.execute(
            """
        INSERT INTO chantiers (
            id_chantier, departement, gestionnaire, intitule, secteur, adresse,
            province, distance, client, type_client, montant, delai_execution,
            date_marche_gagne, ordre_commencer, rp_demandee, rp_realisee,
            rd_a_demander, rd_realisee, rp_statut, rd_statut,
            n_cautionnement, organisme, montant_cautionnement, document_rp, document_rd,
            prix_de_revient
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
            (
                id_c,
                _safe(g(row, "dep")),
                _safe(g(row, "gest")),
                _safe(g(row, "intitule")),
                _safe(g(row, "secteur")),
                _safe(g(row, "adresse")),
                _safe(g(row, "province")),
                _num(g(row, "distance")),
                _safe(g(row, "client")),
                _safe(g(row, "type_client")),
                _num(g(row, "montant")),
                _num(g(row, "delai")),
                _safe(g(row, "date_marche")),
                _safe(g(row, "ordre")),
                _safe(g(row, "rp_dem")),
                _safe(g(row, "rp_real")),
                _safe(g(row, "rd_dem")),
                _safe(g(row, "rd_real")),
                _safe(g(row, "rp_statut")),
                _safe(g(row, "rd_statut")),
                _safe(g(row, "n_caut")),
                _safe(g(row, "organisme")),
                _num(g(row, "mont_caut")),
                _safe(g(row, "doc_rp")),
                _safe(g(row, "doc_rd")),
                _num(g(row, "pr")),
            ),
        )
        imported += 1

    conn.commit()
    return imported, skipped, col


# ── Import décomptes ─────────────────────────────────────────────────────────
def import_decomptes(
    source,
    sheet_name,
    conn,
    header_row=1,
    first_data_row=2,
    col_id=None,
    col_montant=None,
    col_delai=None,
    col_num=None,
    col_intitule=None,
    col_accepte=None,
    overwrite=False,
):
    wb = open_workbook(source)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    wb.close()

    if not rows:
        return 0, {}

    headers_norm = [_normalize(str(c) if c is not None else "") for c in rows[0]]

    if col_id is None:
        col_id = _find_col(headers_norm, "n° chantier", "num chantier", "id chantier")
        if col_id is None:
            col_id = next(
                (
                    i
                    for i, h in enumerate(headers_norm)
                    if h.strip() in ("n° chantier", "num chantier", "chantier")
                ),
                None,
            )
        if col_id is None:
            col_id = 0
    if col_montant is None:
        # Priorité : montant approuvé > montant décompte > montant
        col_montant = _find_col(headers_norm, "montant approuve", "montant approve")
        if col_montant is None:
            col_montant = _find_col(headers_norm, "montant decompte", "montant dc")
        if col_montant is None:
            col_montant = _find_col(headers_norm, "montant")
    if col_delai is None:
        # Priorité : délai approuvé > délai complémentaire > délai
        col_delai = _find_col(headers_norm, "delai approuve", "delai approve")
        if col_delai is None:
            col_delai = _find_col(headers_norm, "delai complementaire", "delai comp")
        if col_delai is None:
            col_delai = _find_col(headers_norm, "delai")
    if col_num is None:
        col_num = _find_col(
            headers_norm, "n° dc", "n dc", "num dc", "n° decompte", "n decompte"
        )
        if col_num is None:
            col_num = _find_col(headers_norm, "denomination")
    if col_intitule is None:
        col_intitule = _find_col(
            headers_norm, "denomination", "intitul", "libelle", "objet"
        )
    if col_accepte is None:
        col_accepte = _find_col(
            headers_norm, "approuve", "approuv", "accepte", "accept"
        )

    def g(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _accepte(val):
        """Retourne 1 si accepté, 0 sinon, None si colonne absente."""
        if val is None:
            return None
        try:
            return 1 if float(str(val).strip()) == 1.0 else 0
        except (ValueError, TypeError):
            s = str(val).strip().lower()
            return 1 if s in ("1", "oui", "yes", "true", "x") else 0

    if overwrite:
        conn.execute("DELETE FROM decomptes")
        conn.commit()

    imported = 0
    for row in rows[first_data_row - header_row :]:
        id_c = _safe(g(row, col_id))
        if not id_c:
            continue
        accepte_val = _accepte(g(row, col_accepte)) if col_accepte is not None else None
        conn.execute(
            "INSERT INTO decomptes (id_chantier, n_decompte, intitule, montant, delai_complementaire, statut_accepte) "
            "VALUES (?,?,?,?,?,?)",
            (
                id_c,
                _safe(g(row, col_num)),
                _safe(g(row, col_intitule)),
                _num(g(row, col_montant)),
                _num(g(row, col_delai)),
                accepte_val,
            ),
        )
        imported += 1

    conn.commit()
    return imported, {
        "col_id": col_id,
        "col_montant": col_montant,
        "col_delai": col_delai,
        "col_num": col_num,
        "col_accepte": col_accepte,
    }


# ── Import états d'avancement ─────────────────────────────────────────────────
def import_ea(
    source,
    sheet_name,
    conn,
    header_row=1,
    first_data_row=2,
    col_id=None,
    col_montant=None,
    overwrite=False,
):
    """
    Import états d'avancement.
    col_id      : index colonne num chantier (None = auto-détecté)
    col_montant : index colonne montant facturé (None = auto-détecté)
    """
    wb = open_workbook(source)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    wb.close()

    if not rows:
        return 0, "Feuille vide"

    headers_norm = [_normalize(str(c) if c is not None else "") for c in rows[0]]

    if col_id is None:
        col_id = _find_col(headers_norm, "num chantier", "id chantier", "chantier") or 0
    _EA_MONTANT_FALLBACK_COL = 21  # colonne V dans le format EA BEG standard
    if col_montant is None:
        col_montant = (
            _find_col(headers_norm, "revision", "facture", "montant facture")
            or _find_col(headers_norm, "montant")
            or _EA_MONTANT_FALLBACK_COL
        )

    col_lot = _find_col(headers_norm, "lot")
    col_gest = _find_col(headers_norm, "gestionnaire", "gest")
    col_nom = _find_col(headers_norm, "nom chantier", "nom de chantier")
    col_date = _find_col(headers_norm, "date envoi", "date d'envoi")
    col_nea = _find_col(headers_norm, "n° ea", "n ea", "numero ea")

    def g(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    if overwrite:
        conn.execute("DELETE FROM etats_avancement")
        conn.commit()

    imported = 0
    for row in rows[first_data_row - header_row :]:
        num_c = _safe(g(row, col_id))
        if not num_c:
            continue
        conn.execute(
            """
        INSERT INTO etats_avancement
            (num_chantier, lot, gestionnaire, nom_chantier, date_envoi, n_ea, montant_facture)
        VALUES (?,?,?,?,?,?,?)
        """,
            (
                num_c,
                _safe(g(row, col_lot)),
                _safe(g(row, col_gest)),
                _safe(g(row, col_nom)),
                _safe(g(row, col_date)),
                _safe(g(row, col_nea)),
                _num(g(row, col_montant)),
            ),
        )
        imported += 1

    conn.commit()
    return imported, {"col_id": col_id, "col_montant": col_montant}
